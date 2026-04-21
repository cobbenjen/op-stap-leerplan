import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Zet LP-wiskunde.xlsx om naar data.json met kolommen F t/m K."
    )
    parser.add_argument(
        "--input",
        default="LP-wiskunde.xlsx",
        help="Pad naar het Excel-bestand (standaard: LP-wiskunde.xlsx).",
    )
    parser.add_argument(
        "--output",
        default="data.json",
        help="Pad naar het JSON-bestand (standaard: data.json).",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optioneel: naam van het werkblad. Zonder deze optie wordt het eerste werkblad gebruikt.",
    )
    return parser.parse_args()


def normalize(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def excel_naar_json(input_path: Path, output_path: Path, sheet_name: str | None = None) -> None:
    wb = load_workbook(filename=input_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    records = []
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=11, values_only=True):
        fase, domein, subdomein, col_i, col_j, col_k = map(normalize, row)

        if all(v is None for v in (fase, domein, subdomein, col_i, col_j, col_k)):
            continue

        records.append(
            {
                "fase": fase,
                "domein": domein,
                "subdomein": subdomein,
                "Cluster": col_i,
                "Leerplandoel": col_j,
                "Voorbeelden": col_k,
            }
        )

    output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Klaar: {len(records)} records geschreven naar {output_path}")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Inputbestand niet gevonden: {input_path}")

    excel_naar_json(input_path, output_path, args.sheet)


if __name__ == "__main__":
    main()
